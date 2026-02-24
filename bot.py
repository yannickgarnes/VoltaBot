import os
import json
import tempfile
import time
import re
import shutil
import glob
import traceback
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# ============================================================
# CONFIGURACIÓN - Railway usa variables de entorno
# ============================================================
EXCEL_PATH = "/tmp/FIFA_VOLTA.xlsx"
URL = "https://www.bet365.es/#/IP/B1"
GSHEET_ID = os.environ.get("GSHEET_ID", "")

def get_creds_path():
    creds_str = os.environ.get("GOOGLE_CREDS_JSON", "")
    if not creds_str:
        print("❌ No se encontró GOOGLE_CREDS_JSON en variables de entorno")
        return None
    try:
        creds_dict = json.loads(creds_str)
        tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)
        json.dump(creds_dict, tmp)
        tmp.close()
        return tmp.name
    except Exception as e:
        print(f"❌ Error procesando credenciales: {e}")
        return None

CREDS_JSON = get_creds_path()
partidos_monitoreados = {}

# ============================================================
# FUNCIONES DE EXCEL LOCAL (temporal en /tmp)
# ============================================================

def preparar_excel():
    try:
        expected = [
            'EQUIPO 1', 'EQUIPO 2', '1P 1', '1P 2', '2P 1', '2P 2', 'TOTAL',
            'CUOTA AMBOS MARCAN 1 PARTE', 'AMBOS MARCAN'
        ]
        if not os.path.exists(EXCEL_PATH):
            df = pd.DataFrame(columns=expected)
            df.to_excel(EXCEL_PATH, index=False)
            print(f"📁 Excel temporal creado en {EXCEL_PATH}")
        else:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for col in expected:
                if col not in headers:
                    ws.cell(row=1, column=ws.max_column + 1, value=col)
            wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"❌ Error Excel local: {e}")

# ============================================================
# GOOGLE SHEETS - Actualización en tiempo real
# ============================================================

def guardar_en_gsheet(datos_fila, ambos_1p, ambos_partido):
    try:
        if not CREDS_JSON or not GSHEET_ID:
            print("⚠️ Sin credenciales o GSHEET_ID, saltando Google Sheets")
            return

        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive.file",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDS_JSON, scope)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(GSHEET_ID).sheet1

        eq1 = datos_fila['eq1']
        eq2 = datos_fila['eq2']
        g1p1 = datos_fila['g1p1']
        g1p2 = datos_fila['g1p2']
        g2p1 = datos_fila['g2p1']
        g2p2 = datos_fila['g2p2']
        total = f"{g1p1 + g2p1}-{g1p2 + g2p2}"

        nueva_fila = [eq1, eq2, g1p1, g1p2, g2p1, g2p2, total, "", ""]
        sheet.append_row(nueva_fila)

        rows = sheet.get_all_values()
        last_row_idx = len(rows)

        color_verde = {"red": 0.0, "green": 0.9, "blue": 0.0}
        color_rojo  = {"red": 1.0, "green": 0.0, "blue": 0.0}

        sheet.format(f"H{last_row_idx}", {"backgroundColor": color_verde if ambos_1p else color_rojo})
        sheet.format(f"I{last_row_idx}", {"backgroundColor": color_verde if ambos_partido else color_rojo})

        print(f"📊 [GSHEETS] ✅ Actualizado: {eq1} vs {eq2} | 1ªP: {g1p1}-{g1p2} | 2ªP: {g2p1}-{g2p2} | Total: {total}")

    except Exception as e:
        print(f"❌ [GSHEETS] Error: {e}")
        traceback.print_exc()

def guardar_resultado(datos):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        headers = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
        row = ws.max_row + 1

        m1 = re.search(r'\((.*?)\)', datos['eq1'])
        eq1 = m1.group(1).strip().upper() if m1 else datos['eq1'].strip().upper()
        m2 = re.search(r'\((.*?)\)', datos['eq2'])
        eq2 = m2.group(1).strip().upper() if m2 else datos['eq2'].strip().upper()

        datos_limpios = datos.copy()
        datos_limpios['eq1'] = eq1
        datos_limpios['eq2'] = eq2

        ws.cell(row=row, column=headers['EQUIPO 1'], value=eq1)
        ws.cell(row=row, column=headers['EQUIPO 2'], value=eq2)
        ws.cell(row=row, column=headers['1P 1'], value=datos['g1p1'])
        ws.cell(row=row, column=headers['1P 2'], value=datos['g1p2'])
        ws.cell(row=row, column=headers['2P 1'], value=datos['g2p1'])
        ws.cell(row=row, column=headers['2P 2'], value=datos['g2p2'])

        total_1 = datos['g1p1'] + datos['g2p1']
        total_2 = datos['g1p2'] + datos['g2p2']
        ws.cell(row=row, column=headers['TOTAL'], value=f"{total_1}-{total_2}")

        ambos_1p = datos['g1p1'] > 0 and datos['g1p2'] > 0
        ambos_partido = total_1 > 0 and total_2 > 0

        cell_h = ws.cell(row=row, column=headers['CUOTA AMBOS MARCAN 1 PARTE'], value="")
        cell_h.fill = PatternFill(start_color="00FF00" if ambos_1p else "FF0000", fill_type="solid")

        cell_i = ws.cell(row=row, column=headers['AMBOS MARCAN'], value="")
        cell_i.fill = PatternFill(start_color="00FF00" if ambos_partido else "FF0000", fill_type="solid")

        wb.save(EXCEL_PATH)
        print(f"✅ EXCEL OK: {eq1} vs {eq2} | Final: {total_1}-{total_2}")

        # Subir a Google Sheets
        guardar_en_gsheet(datos_limpios, ambos_1p, ambos_partido)

    except Exception as e:
        print(f"❌ Error guardando resultado: {e}")
        traceback.print_exc()

# ============================================================
# BUCLE PRINCIPAL - 24/7 con auto-reinicio
# ============================================================

def ejecutar_bot():
    print("📢 Iniciando preparar_excel()...")
    preparar_excel()
    print("🚀 BOT VOLTA INICIADO EN RAILWAY - 24/7")
    print(f"🕐 {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"📋 Google Sheet ID: {GSHEET_ID}")

    # --- DETECCIÓN DE CHROMIUM ---
    chrome_path = os.environ.get("CHROMIUM_PATH")
    
    if not chrome_path:
        búsqueda = ["chromium", "chromium-browser", "google-chrome", "google-chrome-stable"]
        for b in búsqueda:
            chrome_path = shutil.which(b)
            if chrome_path: break
    
    if not chrome_path:
        vias = [
            "/usr/bin/chromium", 
            "/usr/bin/chromium-browser", 
            "/usr/bin/google-chrome",
            "/usr/bin/google-chrome-stable",
            "/nix/var/nix/profiles/default/bin/chromium",
            "/nix/var/nix/profiles/default/bin/google-chrome"
        ]
        for v in vias:
            if os.path.exists(v):
                chrome_path = v
                break
    
    if not chrome_path:
        print("🔍 Buscando en /nix/store... (puede tardar un poco)")
        posibles = glob.glob("/nix/store/*/bin/chromium") + glob.glob("/nix/store/*/bin/google-chrome")
        if posibles:
            chrome_path = posibles[0]

    if not chrome_path:
        print("❌ ERROR CRÍTICO: No se encuentra Chromium/Chrome en el sistema.")
        print(f"PATH actual: {os.environ.get('PATH')}")
        try:
            print("Contenido de /usr/bin (primeros 20):", os.listdir("/usr/bin")[:20])
        except:
            pass
        return

    print(f"✅ Navegador detectado en: {chrome_path}")

    while True:
        driver = None
        try:
            print(f"\n🔄 [{datetime.now().strftime('%H:%M:%S')}] Configurando ChromeOptions...")

            options = uc.ChromeOptions()
            options.add_argument("--headless=new")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-blink-features=AutomationControlled")
            
            print(f"🛠️ Llamando a uc.Chrome con path: {chrome_path}")
            # Agregamos use_subprocess=True que ayuda en entornos Docker/Linux
            driver = uc.Chrome(options=options, browser_executable_path=chrome_path, use_subprocess=True)
            print("🌐 Navegador iniciado correctamente. Cargando URL...")
            
            wait = WebDriverWait(driver, 20)
            driver.get(URL)
            print(f"📍 URL cargada: {driver.current_url}")
            time.sleep(10)

            # Aceptar cookies si aparecen
            try:
                cookies = driver.find_element(By.XPATH, "//div[contains(text(), 'Aceptar')]")
                cookies.click()
                time.sleep(2)
                print("🍪 Cookies aceptadas")
            except:
                pass

            fail_count = 0

            while True:
                try:
                    if URL not in driver.current_url:
                        driver.get(URL)
                        time.sleep(5)

                    comp_elements = driver.find_elements(By.CLASS_NAME, "ovm-Competition")
                    volta_section = None
                    for comp in comp_elements:
                        try:
                            if "Battle Volta" in comp.text:
                                volta_section = comp
                                break
                        except:
                            continue

                    en_pantalla = set()

                    if volta_section:
                        fail_count = 0
                        try:
                            fixtures = volta_section.find_elements(By.CLASS_NAME, "ovm-Fixture")
                        except:
                            fixtures = []

                        for fixture in fixtures:
                            try:
                                names = fixture.find_elements(By.CLASS_NAME, "ovm-FixtureDetailsTwoWay_TeamName")
                                if len(names) < 2:
                                    continue
                                eq_raw1 = names[0].text.strip()
                                eq_raw2 = names[1].text.strip()
                                id_match = f"{eq_raw1} vs {eq_raw2}"
                                en_pantalla.add(id_match)

                                s1_el = fixture.find_element(By.CLASS_NAME, "ovm-StandardScoresSoccer_TeamOne")
                                s2_el = fixture.find_element(By.CLASS_NAME, "ovm-StandardScoresSoccer_TeamTwo")
                                s1, s2 = int(s1_el.text.strip()), int(s2_el.text.strip())

                                timer_el = fixture.find_element(By.CLASS_NAME, "ovm-FixtureDetailsTwoWay_Timer")
                                timer_str = timer_el.text.strip()

                                t_match = re.search(r'(\d{2}):(\d{2})', timer_str)
                                minutos = int(t_match.group(1)) if t_match else 0
                                segundos = int(t_match.group(2)) if t_match else 0

                                if id_match not in partidos_monitoreados:
                                    print(f"🆕 [{datetime.now().strftime('%H:%M:%S')}] Detectado: {id_match} ({timer_str})")
                                    partidos_monitoreados[id_match] = {
                                        "eq1": eq_raw1, "eq2": eq_raw2,
                                        "estado": "jugando_1p",
                                        "g1p1": 0, "g1p2": 0,
                                        "g2p1": 0, "g2p2": 0,
                                        "cuota_1p": 1.0, "cuota_partido": 1.0,
                                        "ultimo_s1_visto": s1, "ultimo_s2_visto": s2,
                                        "marcador_pre_3min": (s1, s2),
                                        "ultimo_min": minutos
                                    }

                                p = partidos_monitoreados[id_match]
                                p["ultimo_s1_visto"] = s1
                                p["ultimo_s2_visto"] = s2
                                p["ultimo_min"] = minutos

                                if minutos < 3:
                                    p["marcador_pre_3min"] = (s1, s2)

                                if p["estado"] == "jugando_1p":
                                    if "Descanso" in timer_str or "HT" in timer_str or (minutos == 3 and segundos <= 5):
                                        print(f"🌘 MEDIA PARTE ({timer_str}): {id_match} -> {s1}-{s2}")
                                        p.update({"g1p1": s1, "g1p2": s2, "estado": "jugando_2p"})
                                    elif minutos >= 3:
                                        g1_prev, g2_prev = p["marcador_pre_3min"]
                                        print(f"🌘 MEDIA PARTE RECUPERADA ({timer_str}): {id_match} -> {g1_prev}-{g2_prev}")
                                        p.update({"g1p1": g1_prev, "g1p2": g2_prev, "estado": "jugando_2p"})

                                elif p["estado"] == "jugando_2p":
                                    if minutos >= 6 or "Finalizado" in timer_str or "FT" in timer_str:
                                        print(f"🏁 FINAL ({timer_str}): {id_match} -> {s1}-{s2}")
                                        p.update({
                                            "g2p1": s1 - p["g1p1"],
                                            "g2p2": s2 - p["g1p2"],
                                            "estado": "finalizado"
                                        })
                                        guardar_resultado(p)

                            except Exception as e_row:
                                if "stale" in str(e_row).lower():
                                    continue
                                print(f"⚠️ Error partido: {e_row}")
                                continue

                    else:
                        fail_count += 1
                        print(f"🔍 [{datetime.now().strftime('%H:%M:%S')}] Sin sección Volta visible... ({fail_count})")
                        if fail_count > 10:
                            driver.execute_script("window.scrollBy(0, 500);")
                            fail_count = 0

                    # Limpieza y rescate de partidos desaparecidos
                    borrar_lista = []
                    for mid, p in partidos_monitoreados.items():
                        if mid not in en_pantalla and p["estado"] not in ("finalizado", "ignorado"):
                            if p["ultimo_min"] >= 5:
                                print(f"🚨 Rescatando partido desaparecido: {mid}")
                                p.update({
                                    "g2p1": p["ultimo_s1_visto"] - p["g1p1"],
                                    "g2p2": p["ultimo_s2_visto"] - p["g1p2"],
                                    "estado": "finalizado"
                                })
                                guardar_resultado(p)
                            borrar_lista.append(mid)

                    if len(partidos_monitoreados) > 30:
                        for m in borrar_lista:
                            del partidos_monitoreados[m]

                    time.sleep(3)

                except Exception as e_inner:
                    print(f"⚠️ Error bucle interno: {e_inner}")
                    if "stale" in str(e_inner).lower():
                        driver.get(URL)
                    time.sleep(5)

        except Exception as e_outer:
            print(f"❌ FALLO CRÍTICO: {e_outer}")
            print(f"🔁 Reiniciando en 15 segundos...")
            traceback.print_exc()
            if driver:
                try:
                    driver.quit()
                except:
                    pass
            time.sleep(15)

if __name__ == "__main__":
    ejecutar_bot()
